"""
Indian bank statement -> grouped Deposits / Withdrawals workbook.

Architecture:
  detect_profile(path)  -> Profile (icici | axis | hdfc | sbi | kotak | generic)
  Profile.parse(path)   -> DataFrame[txn_date, remarks, withdrawal, deposit, balance, counterparty]
  consolidate(...)      -> merge near-duplicate counterparties
  write_workbook(...)   -> 2-sheet xlsx grouped by counterparty, threshold >=2

CLI:
  python sort_statement.py <statement_file> [output.xlsx]

Adding a new bank: write a new Profile subclass with detect() + parse_columns()
+ extract_counterparty() and register it in PROFILES.
"""

from __future__ import annotations

import sys

if sys.version_info < (3, 8):
    print(
        f"Minsorterbank requires Python 3.8 or later.\n"
        f"You are running Python {sys.version}.\n"
        f"Please install a newer Python from https://www.python.org/downloads/"
    )
    sys.exit(1)

import csv
import io
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# Shared normalization / consolidation helpers
# ============================================================

_STOPWORDS = {
    "LTD", "LIMITED", "PVT", "PRIVATE", "CO", "COMPANY", "INC", "CORP",
    "INDIA", "THE", "OF", "AND", "PAYT", "BANK",
    "MR", "MRS", "MS", "DR", "SHRI", "SMT",
    "PAYMENT", "PAYMENTS", "TRANSFER", "ONLINE", "UPI",
}

# Generic free-text notes that aren't real counterparties.
_GENERIC_NOTES = {
    "UPISCANQR", "UPISENDMONEY", "UPI", "UPI TRANSACTION", "UPI PAYMENT",
    "PAYMENT", "COLLECT TRANSAC", "PAY VIA RAZORPA", "PAY TO BHARATPE",
    "SOME NOTE", "GENERATING DYNA", "CCA", "ONLINE", "TRANSFER",
    "OTHERS", "OTHPOS", "POS", "ATM CASH", "ATM WDL", "ATMWDL",
    "PAY TO", "PAY", "PAYTO", "PAYVIA", "COLLEC", "REFUND",
    "SELF TRANSFER", "FD THROUGH NET", "INB", "P2M", "P2A",
}

# Gateway VPAs are aggregators, not merchants.
_GATEWAY_VPA = {
    "PAYTM", "RAZORPAY", "BHARATPE", "PHONEPE", "GPAY", "GOOGLEPAY",
    "OKAXIS", "OKHDFCBANK", "OKSBI", "OKICICI", "AMAZONPAY",
}

# Merchant family aliases — collapse rail/gateway variants.
_MERCHANT_ALIASES = [
    (("ZOMATO", "PAYZOMATO", "ZOMATO ONLINE", "ZOMATO1PAYTM"), "ZOMATO"),
    (("SWIGGY", "BUNDL TECHNOLOGIES"), "SWIGGY"),
    (("BIGBASKET",), "BIGBASKET"),
    (("TATAPAY",), "TATA PAY (utilities)"),
    (("AMAZONPAY", "AMAZON PAY"), "AMAZON PAY"),
    (("DOMINOS",), "DOMINOS"),
    (("UBER",), "UBER"),
    (("DUNZO",), "DUNZO"),
    (("FRANKLIN", "FRANKLINTEMP"), "FRANKLIN TEMPELTON MUTUAL FUND"),
    (("INVESTEASYRMF",), "ICICI INVESTEASY"),
    (("APOLLO",), "APOLLO SPECIALTY HOSPITALS PVT LTD"),
    (("KIDS CLINIC", "KIDSCLINIC"), "KIDS CLINIC INDIA LIMITED"),
    (("AMERCIAN E", "AMEX", "AMERICAN EXPRESS", "AMERICANEX"), "AMERICAN EXPRESS"),
    (("ICICI BANK CREDIT CA",), "ICICI CREDIT CARD"),
    (("HDFCVI",), "HDFC CREDIT CARD"),
    (("BBMP",), "BBMP"),
    (("RAZPGROWW",), "GROWW (Razorpay)"),
    (("IMPERIAL HOSPIT",), "IMPERIAL HOSPITAL"),
    (("HUZUR TREA", "STATE HUZUR"), "STATE HUZUR TREASURY"),
    (("KOTAK CUSTODY",), "KOTAK CUSTODY (RTGS/NEFT)"),
]


def _norm(s) -> str:
    if not s:
        return ""
    s = re.sub(r"\s+", " ", str(s).strip().upper())
    s = re.sub(r"\s+\d+$", "", s)
    s = s.rstrip(" -_/.,;:")
    s = s.lstrip(" -_/.,;:")
    return s


def _apply_alias(key: str) -> str:
    if not key:
        return key
    up = key.upper()
    for patterns, canonical in _MERCHANT_ALIASES:
        for pat in patterns:
            if pat in up:
                return canonical
    return key


def _clean_vpa(v: str) -> Optional[str]:
    if "@" not in v:
        return None
    h = v.split("@")[0]
    if not h or re.fullmatch(r"\d{8,}", h):
        return None
    h = re.sub(r"\.\d+$", "", h)
    h = re.sub(r"\d+$", "", h)
    n = _norm(h)
    return n if n and len(n) >= 3 else None


def _compressed(s) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def _tokens(key) -> set[str]:
    if not key:
        return set()
    return {t for t in re.split(r"[^A-Z0-9]+", str(key))
            if t and t not in _STOPWORDS and len(t) > 2}


def consolidate(keys: list[str]) -> dict[str, str]:
    """Merge near-duplicate counterparties. Returns {raw: canonical}."""
    clean = sorted({str(k) for k in keys if k and isinstance(k, str) and k.strip()})
    if not clean:
        return {}

    groups: dict[str, list[str]] = defaultdict(list)
    for k in clean:
        toks = sorted(_tokens(k), key=lambda t: (-len(t), t))
        anchor = toks[0] if toks else _compressed(k)[:8] or k
        groups[anchor].append(k)

    group_keys = list(groups.keys())
    parent = {g: g for g in group_keys}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    reps = {g: _compressed(max(members, key=len)) for g, members in groups.items()}

    for i, g1 in enumerate(group_keys):
        r1 = reps[g1]
        if len(r1) < 6:
            continue
        for g2 in group_keys[i + 1:]:
            r2 = reps[g2]
            if len(r2) < 6:
                continue
            short, long_ = (r1, r2) if len(r1) <= len(r2) else (r2, r1)
            if len(short) >= 8 and short in long_:
                union(g1, g2)
                continue
            common = 0
            for c1, c2 in zip(r1, r2):
                if c1 != c2:
                    break
                common += 1
            if common >= 10:
                union(g1, g2)

    merged: dict[str, list[str]] = defaultdict(list)
    for g, members in groups.items():
        merged[find(g)].extend(members)

    return {m: max(members, key=len) for _, members in merged.items() for m in members}


# ============================================================
# Generic counterparty extractor (last-resort fallback for unknown formats)
# ============================================================

def extract_generic(remarks: str) -> Optional[str]:
    """When no bank profile recognised the remark, fall back to this.
    Strategy: pull the longest run of uppercase words (likely a name/merchant)
    or the longest non-numeric token in the remark."""
    r = str(remarks).strip()
    if not r:
        return None
    # Try uppercase-word runs first (e.g. 'MR KUMAR PRABHAS', 'KIDS CLINIC INDIA')
    runs = re.findall(r"\b([A-Z][A-Z &]{2,}(?:\s+[A-Z][A-Z &]*){0,5})\b", r)
    runs = [s.strip() for s in runs if s.strip() and _norm(s) not in _GENERIC_NOTES]
    if runs:
        return _norm(max(runs, key=len))
    # Fall back: longest alpha token
    tokens = re.findall(r"[A-Za-z]{4,}", r)
    tokens = [t for t in tokens if _norm(t) not in _GENERIC_NOTES]
    if tokens:
        return _norm(max(tokens, key=len))
    return None


# ============================================================
# Profiles
# ============================================================

class Profile:
    name = "base"

    @classmethod
    def detect(cls, path: Path, raw_text: str, raw_df: Optional[pd.DataFrame]) -> bool:
        return False

    @classmethod
    def parse(cls, path: Path, header_row: Optional[int] = None) -> pd.DataFrame:
        raise NotImplementedError

    @classmethod
    def extract(cls, remarks: str) -> Optional[str]:
        return extract_generic(remarks)


def _load_raw(path: Path):
    """Return (raw_text_first_8k, raw_df_or_None). Used for detection and parsing.

    Tries (in order): xlrd (.xls BIFF), openpyxl (.xlsx), pd.read_html
    (many bank exports are HTML wrapped in a .xls extension)."""
    raw_text = ""
    try:
        with open(path, "rb") as f:
            raw_text = f.read(8192).decode("utf-8", errors="replace")
    except Exception:
        pass
    raw_df = None
    for engine in ("xlrd", "openpyxl"):
        try:
            raw_df = pd.read_excel(path, sheet_name=0, engine=engine, header=None)
            break
        except Exception:
            continue
    if raw_df is None:
        try:
            tables = pd.read_html(str(path), header=None)
            if tables:
                raw_df = max(tables, key=lambda t: t.shape[0]).reset_index(drop=True)
                raw_df.columns = range(raw_df.shape[1])
        except Exception:
            pass
    return raw_text, raw_df


# ============================================================
# Shared header-driven row parser
# ============================================================

_DATE_TOKENS = ("txn date", "tran date", "transaction date", "value date", "date")
_REMARKS_TOKENS = ("transaction remarks", "narration", "particulars",
                   "description", "remarks")
_DEBIT_TOKENS = ("withdrawal amt", "withdrawal", "debit amt", "debit", "dr")
_CREDIT_TOKENS = ("deposit amt", "deposit", "credit amt", "credit", "cr")
_BALANCE_TOKENS = ("closing balance", "balance", "bal")


def _header_match(needle: str, cell: str) -> bool:
    """Word-boundary needle match. Prevents 'cr' from hitting 'description'."""
    if not cell:
        return False
    return re.search(rf"\b{re.escape(needle)}", cell) is not None


def _find_header_row_any(df: pd.DataFrame) -> Optional[int]:
    """Find the row that looks like a transaction header.

    Strong match: date + remarks + amount tokens all present.
    Weak match (fallback): date + amount tokens present (remarks column is
    inferred later as the widest non-date/non-amount text column).
    """
    weak_match: Optional[int] = None
    for i in range(min(60, len(df))):
        row_lc = [str(v).lower().strip() for v in df.iloc[i].tolist()]
        has_date = any(any(_header_match(t, v) for t in _DATE_TOKENS) for v in row_lc)
        has_rem = any(any(_header_match(t, v) for t in _REMARKS_TOKENS) for v in row_lc)
        has_amt = any(any(_header_match(t, v) for t in _DEBIT_TOKENS + _CREDIT_TOKENS)
                      for v in row_lc)
        if has_date and has_rem and has_amt:
            return i
        if has_date and has_amt and weak_match is None:
            weak_match = i
    return weak_match


def _pick_col(header: list[str], needles: tuple[str, ...],
              avoid: Optional[set[int]] = None) -> Optional[int]:
    """Return the column index whose lowercase header matches one of the needles.

    Iterates needles outer-loop first so the most specific token (e.g. 'txn date')
    wins over a less specific one (e.g. 'date' matching 'Value Date')."""
    avoid = avoid or set()
    for n in needles:
        for idx, h in enumerate(header):
            if idx in avoid:
                continue
            if _header_match(n, h):
                return idx
    return None


class HeaderNotFoundError(RuntimeError):
    """Raised when the header row can't be auto-detected.

    Carries a structured preview of the first rows so the UI can ask the user
    to pick the header row manually.
    """

    def __init__(self, preview: list[dict]):
        self.preview = preview
        super().__init__(
            "Couldn't locate transaction header row. Expected a row containing "
            "a date column (e.g. 'Date', 'Txn Date') and an amount column "
            "(e.g. 'Debit', 'Credit', 'Withdrawal')."
        )


def _classify_columns(body: pd.DataFrame) -> dict[int, str]:
    """Tag each column as 'date', 'amount', 'text', or 'empty' from its data.

    Used as a fallback when header tokens don't match (e.g. non-English headers
    like 'Tarikh' / 'Money Out') so the user-confirmed header row still yields a
    usable parse."""
    classes: dict[int, str] = {}
    for idx in range(body.shape[1]):
        col = body.iloc[:, idx]
        non_null = col[col.notna()].astype(str).str.strip()
        non_null = non_null[non_null.ne("") & non_null.str.lower().ne("nan")]
        if len(non_null) == 0:
            classes[idx] = "empty"
            continue
        cleaned = non_null.str.replace(",", "", regex=False) \
            .str.replace("₹", "", regex=False).str.strip().str.strip('"').str.strip()
        num_ratio = pd.to_numeric(cleaned, errors="coerce").notna().mean()
        # Suppress noisy dateutil warnings while sniffing
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            dt_ratio = pd.to_datetime(non_null, errors="coerce", dayfirst=True).notna().mean()
        if dt_ratio >= 0.7 and dt_ratio >= num_ratio:
            classes[idx] = "date"
        elif num_ratio >= 0.7:
            classes[idx] = "amount"
        else:
            classes[idx] = "text"
    return classes


def _build_preview(raw: pd.DataFrame, limit: int = 15) -> list[dict]:
    preview = []
    for i in range(min(limit, len(raw))):
        cells = [str(v) for v in raw.iloc[i].tolist() if str(v).strip() and str(v) != "nan"]
        if cells:
            preview.append({"index": i, "text": " | ".join(cells)[:200]})
    return preview


def _resolve_columns_from_header(raw: pd.DataFrame, header_row: int) -> dict:
    """Resolve transaction column indices from a header row + the data body.

    Returns ``{"col_date", "col_remarks", "col_withdrawal", "col_deposit",
    "col_balance"}``. Any key may be ``None`` if neither header text nor data
    inference could place a column."""
    header = [str(v).lower().strip() for v in raw.iloc[header_row].tolist()]
    body = raw.iloc[header_row + 1:]

    col_date = _pick_col(header, _DATE_TOKENS)
    col_wd = _pick_col(header, _DEBIT_TOKENS)
    col_dep = _pick_col(header, _CREDIT_TOKENS, avoid={col_wd} if col_wd is not None else None)
    col_bal = _pick_col(header, _BALANCE_TOKENS)
    col_rem = _pick_col(header, _REMARKS_TOKENS)

    classified = _classify_columns(body)
    used = {c for c in (col_date, col_wd, col_dep, col_bal, col_rem) if c is not None}

    if col_date is None:
        for idx, kind in classified.items():
            if idx not in used and kind == "date":
                col_date = idx
                used.add(idx)
                break

    if col_wd is None or col_dep is None:
        amount_cols = [idx for idx, kind in classified.items()
                       if idx not in used and kind == "amount"]
        debit_hints = ("out", "debit", "dr", "withdraw", "paid", "spent")
        credit_hints = ("in", "credit", "cr", "deposit", "received", "in ", "earned")

        def _hint_side(idx):
            h = header[idx] if idx < len(header) else ""
            if any(t in h for t in debit_hints):
                return "wd"
            if any(t in h for t in credit_hints):
                return "dep"
            return None

        unassigned = []
        for idx in amount_cols:
            side = _hint_side(idx)
            if side == "wd" and col_wd is None:
                col_wd = idx
            elif side == "dep" and col_dep is None:
                col_dep = idx
            else:
                unassigned.append(idx)

        for idx in unassigned:
            if col_wd is None:
                col_wd = idx
            elif col_dep is None:
                col_dep = idx
        used = {c for c in (col_date, col_wd, col_dep, col_bal, col_rem) if c is not None}

    if col_rem is None:
        best_idx, best_width = None, 0
        for idx in range(body.shape[1]):
            if idx in used:
                continue
            col_vals = body.iloc[:, idx].astype(str)
            width = col_vals.str.len().mean() if len(col_vals) else 0
            if width > best_width:
                best_width, best_idx = width, idx
        col_rem = best_idx

    return {
        "col_date": col_date,
        "col_remarks": col_rem,
        "col_withdrawal": col_wd,
        "col_deposit": col_dep,
        "col_balance": col_bal,
    }


def _numify(series: pd.Series) -> pd.Series:
    # Strip currency symbols / commas / quotes before to_numeric
    cleaned = series.astype(str).str.replace(",", "", regex=False) \
        .str.replace("₹", "", regex=False).str.strip().str.strip('"').str.strip()
    return pd.to_numeric(cleaned, errors="coerce").fillna(0)


def _build_txn_frame(
    raw: pd.DataFrame,
    data_start_row: int,
    col_date: Optional[int],
    col_remarks: Optional[int],
    col_withdrawal: Optional[int],
    col_deposit: Optional[int],
    col_balance: Optional[int],
    extract_fn,
) -> pd.DataFrame:
    """Slice the body and project it into the canonical txn DataFrame."""
    df = raw.iloc[data_start_row:].copy().reset_index(drop=True)

    def col(i):
        return df.iloc[:, i] if i is not None else pd.Series([None] * len(df))

    out = pd.DataFrame({
        "txn_date": col(col_date),
        "remarks": col(col_remarks),
        "withdrawal": _numify(col(col_withdrawal)).abs(),
        "deposit": _numify(col(col_deposit)).abs(),
        "balance": pd.to_numeric(
            col(col_balance).astype(str).str.replace(",", "", regex=False).str.strip(),
            errors="coerce",
        ),
    })
    out = out.dropna(subset=["remarks"])
    # Drop separator rows (asterisks, header repeats, etc.)
    out = out[~out["remarks"].astype(str).str.match(r"^\s*[\*\-=_]{3,}\s*$")]
    out = out[(out["withdrawal"] > 0) | (out["deposit"] > 0)].reset_index(drop=True)
    out["counterparty"] = out["remarks"].map(extract_fn)
    return out


def parse_with_explicit_columns(
    raw: pd.DataFrame,
    data_start_row: int,
    col_date: int,
    col_remarks: int,
    col_withdrawal: int,
    col_deposit: int,
    col_balance: Optional[int],
    extract_fn,
) -> pd.DataFrame:
    """Parse a transaction DataFrame using user-supplied column indices.

    Skips header auto-detection and column inference entirely. ``col_balance``
    may be ``None`` to mean "no balance column"; the other four are required.
    """
    if raw is None or raw.shape[1] == 0:
        raise ValueError("Could not read any tabular data from the file.")
    n_rows, n_cols = raw.shape
    if data_start_row < 0 or data_start_row >= n_rows:
        raise ValueError(
            f"Data start row {data_start_row} is out of range (file has {n_rows} rows)."
        )
    required = {
        "col_date": col_date,
        "col_remarks": col_remarks,
        "col_withdrawal": col_withdrawal,
        "col_deposit": col_deposit,
    }
    for name, val in required.items():
        if val is None or val < 0 or val >= n_cols:
            raise ValueError(
                f"{name}={val} is out of range (file has {n_cols} columns)."
            )
    if col_balance is not None and (col_balance < 0 or col_balance >= n_cols):
        raise ValueError(
            f"col_balance={col_balance} is out of range (file has {n_cols} columns)."
        )
    return _build_txn_frame(
        raw, data_start_row,
        col_date, col_remarks, col_withdrawal, col_deposit, col_balance,
        extract_fn,
    )


def build_preview(raw: pd.DataFrame, row_limit: int = 0) -> dict:
    """Return preview rows + auto-detected header/column suggestions.

    Used by the ``/api/preview`` endpoint to populate the column-mapping UI.
    Suggestions may be ``None`` when auto-detection can't decide -- the UI
    will just leave those columns as "Ignore" for the user to set.

    ``row_limit=0`` means return all rows (the default).
    """
    if raw is None or raw.shape[1] == 0:
        return {
            "rows": [],
            "n_cols": 0,
            "suggested_header_row": None,
            "suggested_columns": {
                "col_date": None, "col_remarks": None,
                "col_withdrawal": None, "col_deposit": None, "col_balance": None,
            },
        }
    n_rows, n_cols = raw.shape
    limit = min(row_limit, n_rows) if row_limit > 0 else n_rows
    rows: list[list[str]] = []
    for i in range(limit):
        row = []
        for v in raw.iloc[i].tolist():
            s = "" if v is None else str(v)
            if s.lower() == "nan":
                s = ""
            row.append(s)
        rows.append(row)

    header_row = _find_header_row_any(raw)
    if header_row is not None:
        cols = _resolve_columns_from_header(raw, header_row)
    else:
        cols = {
            "col_date": None, "col_remarks": None,
            "col_withdrawal": None, "col_deposit": None, "col_balance": None,
        }
    return {
        "rows": rows,
        "n_cols": n_cols,
        "suggested_header_row": header_row,
        "suggested_columns": cols,
    }


def _parse_from_df(
    raw: pd.DataFrame,
    extract_fn,
    header_row: Optional[int] = None,
) -> pd.DataFrame:
    """Parse a transaction DataFrame by locating columns from header text.

    Used by every bank profile so detection format (xls / xlsx / html) doesn't
    matter — we only care about header names. Pass ``header_row`` to skip
    auto-detection and use that row index as the header."""
    if header_row is not None:
        if header_row < 0 or header_row >= len(raw):
            raise RuntimeError(
                f"Header row {header_row} is out of range (file has {len(raw)} rows)."
            )
        hdr = header_row
    else:
        hdr = _find_header_row_any(raw)
        if hdr is None:
            raise HeaderNotFoundError(_build_preview(raw))

    cols = _resolve_columns_from_header(raw, hdr)
    return _build_txn_frame(
        raw, hdr + 1,
        cols["col_date"], cols["col_remarks"],
        cols["col_withdrawal"], cols["col_deposit"], cols["col_balance"],
        extract_fn,
    )


def _find_header_row(df: pd.DataFrame, *needles: str) -> Optional[int]:
    """Return first row whose joined string contains ALL needles (lowercase substring match)."""
    needles_lc = [n.lower() for n in needles]
    for i in range(min(50, len(df))):
        row_str = " | ".join(str(v).lower() for v in df.iloc[i].tolist())
        if all(n in row_str for n in needles_lc):
            return i
    return None


# ---------- ICICI ----------

class ICICIProfile(Profile):
    name = "icici"

    @classmethod
    def detect(cls, path, raw_text, raw_df):
        text_up = raw_text.upper()
        if "ICICI" in text_up and "TRANSACTION REMARKS" in text_up:
            return True
        if raw_df is None:
            return False
        hdr = _find_header_row(raw_df, "transaction remarks")
        return hdr is not None

    @classmethod
    def parse(cls, path, header_row=None):
        _, raw = _load_raw(path)
        if raw is None:
            raise RuntimeError("ICICI: couldn't load file as a table")
        return _parse_from_df(raw, cls.extract, header_row=header_row)

    @classmethod
    def extract(cls, remarks):
        return _icici_extract(remarks)


def _icici_extract(remarks):
    r = str(remarks).strip()
    if not r:
        return None
    raw = _icici_extract_raw(r)
    if raw is None:
        return None
    if _norm(raw) in _GENERIC_NOTES:
        return None
    return _apply_alias(raw)


def _icici_extract_raw(r):
    # UPI/[<ref>/]<note>/<vpa>/<bank>/<txnid>  OR  UPI/<vpa>/<note>/<bank>/...
    if r.startswith("UPI/"):
        parts = [p.strip() for p in r.split("/")[1:]]
        vpas = [p for p in parts if "@" in p]
        for v in vpas:
            cleaned = _clean_vpa(v)
            if cleaned and cleaned not in _GATEWAY_VPA:
                if "." in cleaned:
                    head, _, tail = cleaned.partition(".")
                    if head in _GATEWAY_VPA and tail:
                        return tail
                return cleaned
        BANK_SUFFIXES = ("BANK", "BANK LTD", "BANK LIMITED", "BANK OF I", "BANK OF INDIA")
        for p in parts:
            if not p or p == "-" or p.isdigit() or "@" in p:
                continue
            up = p.upper()
            if any(s in up for s in BANK_SUFFIXES):
                continue
            if len(p) >= 10 and any(c.isdigit() for c in p) and " " not in p:
                continue
            n = _norm(p)
            if n and n not in _GENERIC_NOTES and len(n) >= 4:
                return n
        for v in vpas:
            cleaned = _clean_vpa(v)
            if cleaned:
                return cleaned
        return None

    if r.startswith("NEFT-") or r.startswith("NEFT/"):
        body = r[5:]
        parts = re.split(r"[-/]", body, maxsplit=2)
        if len(parts) >= 2:
            return _norm(parts[1]) or None
        return None

    if r.startswith("RTGS"):
        body = r[5:] if r.startswith(("RTGS-", "RTGS/")) else r[4:]
        parts = re.split(r"[-/]", body, maxsplit=2)
        if len(parts) >= 2:
            return _norm(parts[1]) or None
        return None

    if r.startswith("BIL/"):
        BIL_SKIP = {"ONL", "INFT", "CCA", "BBPS", "ECU", "BIL"}
        for p in r.split("/")[1:]:
            p2 = p.strip()
            if not p2 or p2 in BIL_SKIP or p2.isdigit():
                continue
            if re.fullmatch(r"[A-Z]{2,4}\d{4,}", p2):
                continue
            return _norm(p2) or None
        return None

    if r.startswith("CMS/"):
        parts = r.split("/")
        if len(parts) >= 3:
            return _norm(parts[2]) or None
        return None

    if r.startswith("ACH/"):
        sub = r.split("/")[1] if "/" in r else ""
        sub_parts = sub.split("-")
        if len(sub_parts) >= 2:
            return _norm(sub_parts[1]) or None
        return _norm(sub) or None

    if r.startswith("CLG/"):
        parts = r.split("/")
        if len(parts) >= 2:
            return _norm(parts[1]) or None
        return None

    if r.startswith("VPS/"):
        parts = r.split("/")
        if len(parts) >= 2:
            return _norm(parts[1]) or None
        return None

    if r.startswith("MMT/") or r.startswith("IMPS/"):
        parts = r.split("/")
        for p in parts[2:]:
            p2 = p.strip()
            if p2 and not p2.isdigit():
                return _norm(p2) or None
        return None

    if r.startswith("GIB/"):
        parts = r.split("/")
        if len(parts) >= 3:
            return _norm(parts[2]) or None
        return None

    if r.startswith("CAM/"):
        if "CASH WDL" in r.upper():
            return "CASH WDL"
        parts = r.split("/")
        if len(parts) >= 3:
            return _norm(parts[2]) or None
        return None

    if r.startswith("NFS/") or r.startswith("ATM/"):
        return "ATM CASH WDL"

    if re.match(r"^\d{6,}:Int\.Pd", r):
        return "SB INTEREST"

    return None


# ---------- Axis ----------

class AxisProfile(Profile):
    name = "axis"

    @classmethod
    def detect(cls, path, raw_text, raw_df):
        text_up = raw_text.upper()
        if "PARTICULARS" in text_up and ("UTIB" in text_up or "AXIS BANK" in text_up):
            return True
        if raw_df is None:
            return False
        hdr = _find_header_row(raw_df, "particulars")
        if hdr is None:
            return False
        head_text = " ".join(str(v) for v in raw_df.iloc[:hdr].values.flatten() if v == v)
        return "UTIB" in head_text or "AXIS" in head_text.upper()

    @classmethod
    def parse(cls, path, header_row=None):
        _, raw = _load_raw(path)
        if raw is None:
            raise RuntimeError("Axis: couldn't load file as a table")
        return _parse_from_df(raw, cls.extract, header_row=header_row)

    @classmethod
    def extract(cls, remarks):
        r = str(remarks).strip()
        if not r:
            return None

        # UPI/P2M|P2A/<ref>/<MERCHANT>/<note>/<bank>
        if r.startswith("UPI/") or r.startswith("IMPS/"):
            parts = [p.strip() for p in r.split("/")]
            # Skip prefix (UPI/IMPS), then P2M/P2A, then ref, then merchant
            if len(parts) >= 4 and parts[1] in ("P2M", "P2A"):
                merchant = parts[3]
                # Sometimes the merchant is followed by trailing dashes/spaces - clean
                n = _norm(merchant)
                if n and n not in _GENERIC_NOTES and len(n) >= 3:
                    return _apply_alias(n)
            # Fallback: any non-empty, non-numeric slot
            for p in parts[2:]:
                if not p or p.isdigit() or "@" in p or p in ("P2M", "P2A"):
                    continue
                n = _norm(p)
                if n and n not in _GENERIC_NOTES and len(n) >= 4 and "BANK" not in n:
                    return _apply_alias(n)
            return None

        # ATM-CASH-AXIS/<atmid>/<...>
        if r.startswith("ATM-CASH"):
            return "ATM CASH WDL"

        # Long format with date/ref like 'IMPERIAL HOSPIT/12042024Imperial117'
        if "/" in r:
            head = r.split("/")[0]
            n = _norm(head)
            if n and n not in _GENERIC_NOTES and len(n) >= 4:
                return _apply_alias(n)

        return _apply_alias(extract_generic(r)) if extract_generic(r) else None


# ---------- HDFC ----------

class HDFCProfile(Profile):
    name = "hdfc"

    @classmethod
    def detect(cls, path, raw_text, raw_df):
        text_up = raw_text.upper()
        # Text-based detection (works for HTML-format .xls files where raw_df is None)
        if "HDFC" in text_up and "NARRATION" in text_up:
            return True
        if raw_df is None:
            return False
        hdr = _find_header_row(raw_df, "narration")
        if hdr is None:
            return False
        head_text = " ".join(str(v) for v in raw_df.iloc[:hdr].values.flatten() if v == v)
        return "HDFC" in head_text.upper()

    @classmethod
    def parse(cls, path, header_row=None):
        _, raw = _load_raw(path)
        if raw is None:
            raise RuntimeError("HDFC: couldn't load file as a table")
        return _parse_from_df(raw, cls.extract, header_row=header_row)

    @classmethod
    def extract(cls, remarks):
        r = str(remarks).strip()
        if not r:
            return None

        # PRIN AND INT AUTO_REDEEM <acc>  -> FD auto-redemption (interest+principal)
        if r.startswith("PRIN AND INT") or "AUTO_REDEEM" in r:
            return "FD AUTO-REDEEM"

        # CBDT/BANK REFERENCE NO:...  -> Income tax (CBDT) payment
        if r.startswith("CBDT/") or "CBDT" in r[:10]:
            return "CBDT (Income Tax)"

        # MC ISSUED - <branch> - <ref> -  - **** <NAME> ****  (manager's cheque)
        m = re.search(r"\*{2,}\s*([^\*]+?)\s*\*{2,}", r)
        if m and "MC ISSUED" in r:
            return _apply_alias(_norm(m.group(1)))

        # SELF - CHQ PAID - <branch>  -> own withdrawal
        if r.startswith("SELF - CHQ PAID") or r.startswith("SELF-CHQ"):
            return "SELF CHEQUE"

        # ACH D- <merchant>-<ref>  or  ACH C-<merchant>-<ref>
        m = re.match(r"^ACH\s+[CD]-\s*([^-]+?)-\d", r)
        if m:
            return _apply_alias(_norm(m.group(1)))

        # NEFT CR-<IFSC>-<NAME>-<rest>  or  RTGS CR-...  or  NEFT DR-...
        m = re.match(r"^(?:NEFT|RTGS|IMPS)\s+(?:CR|DR)-([^-]+)-(.+?)(?:-|$)", r)
        if m:
            name = m.group(2).strip()
            n = _norm(name)
            if n:
                return _apply_alias(n)

        # FT -<NAME> DR - <acc> - <NAME>
        m = re.match(r"^FT\s*-\s*([^\-]+)(?:\s+DR|\s+CR)", r)
        if m:
            return _apply_alias(_norm(m.group(1)))

        # CHQ PAID-MICR CTS-CH-<NAME>
        m = re.match(r"^CHQ PAID-MICR CTS-CH-(.+)$", r)
        if m:
            return _apply_alias(_norm(m.group(1)))

        # IB BILLPAY DR-<merchant>-<acc>
        m = re.match(r"^IB BILLPAY (?:DR|CR)-([^-]+)-", r)
        if m:
            return _apply_alias(_norm(m.group(1)))

        # FD THROUGH NET-<ref>:<NAME>  -> bucket as 'FD'
        if r.startswith("FD THROUGH NET") or r.startswith("FD-OPENED"):
            return "FD (Fixed Deposit creation)"

        # INTEREST CREDIT <acc>
        if r.startswith("INTEREST CREDIT"):
            return "SB INTEREST"

        # POS <ref> <MERCHANT> <city>
        m = re.match(r"^POS\s+\d+\s+(.+?)(?:\s+[A-Z]{3,}\s*$)?$", r)
        if m:
            return _apply_alias(_norm(m.group(1)))

        # UPI-<merchant>-<vpa>-<...>
        if r.startswith("UPI-") or r.startswith("UPI/"):
            parts = re.split(r"[-/]", r)
            for p in parts[1:]:
                p2 = p.strip()
                if not p2 or p2.isdigit() or "@" in p2:
                    continue
                n = _norm(p2)
                if n and n not in _GENERIC_NOTES and len(n) >= 3:
                    return _apply_alias(n)

        # RFX <date><ref> <description>  (forex transactions)
        if r.startswith("RFX ") or re.match(r"^\d{6}RTT\d", r):
            return "FOREX TRANSACTION"

        # NWD-<atm>-<rest>
        if r.startswith("NWD-") or "ATM" in r[:10]:
            return "ATM CASH WDL"

        return _apply_alias(extract_generic(r)) if extract_generic(r) else None


# ---------- SBI ----------

class SBIProfile(Profile):
    name = "sbi"

    @classmethod
    def detect(cls, path, raw_text, raw_df):
        text_up = raw_text.upper()
        if "STATE BANK OF INDIA" in text_up:
            return True
        # Tab-separated text export — be lenient about the IFSC label variant
        if "ACCOUNT NAME" in text_up and ("IFS CODE" in text_up
                                          or "IFSC" in text_up
                                          or "IFS (INDIAN FINANCIAL SYSTEM)" in text_up
                                          or "SBIN0" in text_up):
            return True
        return False

    @classmethod
    def parse(cls, path, header_row=None):
        # If the file loads as a table (HTML-wrapped .xls or real .xls/.xlsx),
        # use the shared header-driven parser.
        _, raw_df = _load_raw(path)
        if raw_df is not None and (
            header_row is not None or _find_header_row_any(raw_df) is not None
        ):
            return _parse_from_df(raw_df, cls.extract, header_row=header_row)

        # Otherwise fall back to the tab-separated text export.
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            text = f.read()
        lines = text.split("\n")
        hdr_idx = None
        for i, ln in enumerate(lines):
            low = ln.lower()
            if "txn date" in low and "description" in low and "debit" in low:
                hdr_idx = i
                break
        if hdr_idx is None:
            raise RuntimeError("SBI: couldn't locate transaction header line")

        header_fields = [h.strip().lower() for h in lines[hdr_idx].split("\t")]

        def find_col(*needles):
            for n in needles:
                for idx, h in enumerate(header_fields):
                    if _header_match(n, h):
                        return idx
            return None

        col_date = find_col(*_DATE_TOKENS)
        col_rem = find_col(*_REMARKS_TOKENS)
        col_wd = find_col(*_DEBIT_TOKENS)
        avoid = {col_wd} if col_wd is not None else set()
        col_dep = None
        for n in _CREDIT_TOKENS:
            for idx, h in enumerate(header_fields):
                if idx in avoid:
                    continue
                if _header_match(n, h):
                    col_dep = idx
                    break
            if col_dep is not None:
                break
        col_bal = find_col(*_BALANCE_TOKENS)

        def _to_num(s):
            s = str(s).strip().strip('"').replace(",", "").strip()
            if not s or s.lower() == "nan":
                return 0.0
            try:
                return float(s)
            except ValueError:
                return 0.0

        rows = []
        for ln in lines[hdr_idx + 1:]:
            if not ln.strip():
                continue
            fields = ln.split("\t")
            if len(fields) <= max(c for c in (col_date, col_rem, col_wd, col_dep) if c is not None):
                continue
            rows.append(fields)

        def get(fields, i):
            return fields[i] if (i is not None and i < len(fields)) else ""

        df = pd.DataFrame({
            "txn_date": [get(r, col_date) for r in rows],
            "remarks": [get(r, col_rem) for r in rows],
            "withdrawal": [abs(_to_num(get(r, col_wd))) for r in rows],
            "deposit": [abs(_to_num(get(r, col_dep))) for r in rows],
            "balance": [_to_num(get(r, col_bal)) for r in rows],
        })
        df = df[(df["withdrawal"] > 0) | (df["deposit"] > 0)].reset_index(drop=True)
        df["counterparty"] = df["remarks"].map(cls.extract)
        return df[["txn_date", "remarks", "withdrawal", "deposit", "balance", "counterparty"]]

    @classmethod
    def extract(cls, remarks):
        r = str(remarks).strip().lstrip().rstrip("-").strip()
        if not r:
            return None

        # TO TRANSFER-INB <merchant>  or  TO TRANSFER-UPI/<...>
        m = re.match(r"^TO TRANSFER-INB\s+(.+?)(?:\s*--)?$", r)
        if m:
            cleaned = re.sub(r"\s+Payments?\s*$", "", m.group(1), flags=re.I).strip()
            return _apply_alias(_norm(cleaned))

        # BY TRANSFER-NEFT*<ifsc>*<ref>*<NAME>  -> NAME slot
        m = re.match(r"^BY TRANSFER-NEFT\*[^*]+\*[^*]+\*(.+?)$", r)
        if m:
            return _apply_alias(_norm(m.group(1)))

        # BY TRANSFER-INB or TO TRANSFER-NEFT (without stars)
        m = re.match(r"^(?:TO|BY) TRANSFER-(?:NEFT|RTGS|IMPS)[*\-](.+?)$", r)
        if m:
            tail = m.group(1)
            # If contains '*', take last *-segment
            if "*" in tail:
                segs = [s for s in tail.split("*") if s.strip()]
                if segs:
                    return _apply_alias(_norm(segs[-1]))
            return _apply_alias(_norm(tail))

        # ATM WDL-ATM CASH <atmid> <loc> -> bucket
        if r.startswith("ATM WDL"):
            return "ATM CASH WDL"

        # TO CLEARING-Chq No. <num> <branch> <NAME>--<num>
        m = re.match(r"^TO CLEARING-Chq No\.\s+\d+\s+\w+\s+(.+?)(?:--\d+)?$", r)
        if m:
            name = re.sub(r"\s*--\s*\d+\s*$", "", m.group(1))
            return _apply_alias(_norm(name))

        # by debit card-OTHPOS<ref> <MERCHANT> <city>
        m = re.match(r"^by debit card-(?:OTHPOS|POS)\d*\s+(.+?)\s+[A-Z]+\s*$", r, re.I)
        if m:
            return _apply_alias(_norm(m.group(1)))

        # TO TRANSFER-UPI/DR/<ref>/<merchant>/<...> or similar
        m = re.match(r"^TO TRANSFER-UPI/[A-Z]+/\d+/([^/]+)/", r)
        if m:
            return _apply_alias(_norm(m.group(1)))

        # BY TRANSFER-Other Bank ... <NAME>
        m = re.match(r"^BY TRANSFER-(.+?)$", r)
        if m:
            tail = m.group(1)
            for sep in ["*", "/"]:
                if sep in tail:
                    segs = [s for s in tail.split(sep) if s.strip() and not s.strip().isdigit()]
                    if segs:
                        return _apply_alias(_norm(segs[-1]))
            return _apply_alias(_norm(tail))

        return _apply_alias(extract_generic(r)) if extract_generic(r) else None


# ---------- Kotak ----------

def _kotak_extract(remarks: str) -> Optional[str]:
    r = str(remarks).strip()
    if not r:
        return None
    up = r.upper()

    rev = up.startswith("REV-") or up.startswith("REV ")
    if rev:
        inner = re.sub(r"^REV[-\s]+", "", r, flags=re.IGNORECASE)
        base = _kotak_extract(inner)
        return f"{base} (REVERSAL)" if base else "REVERSAL"

    if up.startswith("UPI/"):
        parts = [p.strip() for p in r.split("/")]
        if len(parts) >= 2:
            name = parts[1]
            if "@" in name:
                cleaned = _clean_vpa(name)
                if cleaned and cleaned not in _GATEWAY_VPA:
                    return _apply_alias(cleaned)
            idx = 2
            while (len(re.sub(r"[^A-Za-z]", "", name)) < 3
                   and idx < len(parts)
                   and not re.match(r"^\d{6,}", parts[idx])):
                name = f"{name}/{parts[idx]}"
                idx += 1
            n = _norm(name)
            if n and n not in _GENERIC_NOTES and len(n) >= 2:
                return _apply_alias(n)
        return None

    m = re.match(r"^(?:RECD|REED|RECVD)\s*:?\s*IMPS/", up)
    if m:
        parts = [p.strip() for p in r.split("/")]
        for p in parts[2:]:
            if p and not p.isdigit() and "@" not in p:
                n = _norm(p)
                if n and n not in _GENERIC_NOTES and len(n) >= 2:
                    return _apply_alias(n)
        return None

    if up.startswith("SENTIMPS"):
        body = re.sub(r"^SentIMPS", "", r, flags=re.IGNORECASE)
        body = re.sub(r"^\d+", "", body)
        head = body.split("/")[0]
        n = _norm(head)
        if n and n not in _GENERIC_NOTES and len(n) >= 2:
            return _apply_alias(n)
        slots = [s.strip() for s in body.split("/") if s.strip()]
        if slots:
            tail = _norm(slots[-1])
            if tail and tail not in _GENERIC_NOTES and len(tail) >= 3:
                return f"IMPS - {tail}"
        return "IMPS TRANSFER"

    m = re.match(r"^MB\s*:?\s*SENT\s+TO\s+(.+)$", r, flags=re.IGNORECASE)
    if m:
        name = re.split(r"[/]", m.group(1))[0]
        name = re.sub(r"\bTOTAL\b.*$", "", name, flags=re.IGNORECASE)
        n = _norm(name)
        if n and n not in _GENERIC_NOTES:
            return _apply_alias(n)
        return None

    m = re.match(r"^MB\s*:?\s*RECEIVED\s+MONEY\s+FROM\s+(.+)$", r, flags=re.IGNORECASE)
    if m:
        rest = m.group(1).strip()
        if rest.upper().startswith("OWN"):
            return "SELF TRANSFER"
        n = _norm(re.split(r"[/]", rest)[0])
        return _apply_alias(n) if n else "SELF TRANSFER"

    m = re.match(r"^(?:MB|AP|IB)\s*:?\s*BILLPAY\s+(?:FOR\s+)?([A-Za-z][A-Za-z &]+?)\s*\d", r, flags=re.IGNORECASE)
    if m:
        return _apply_alias(_norm(m.group(1)))

    m = re.match(r"^PG\s+\d+\s+([A-Za-z]+)\b", r)
    if m:
        return _apply_alias(_norm(m.group(1)))

    if up.startswith("NEFT"):
        body = re.sub(r"^NEFT[-/ ]*", "", r, flags=re.IGNORECASE)
        toks = body.split()
        if toks and any(c.isdigit() for c in toks[0]):
            toks = toks[1:]
        name = " ".join(toks[:4])
        n = _norm(name)
        if n and n not in _GENERIC_NOTES and len(n) >= 3:
            return _apply_alias(n)
        return None

    m = re.match(r"^IFT[-\s]+([^-]+?)-", r, flags=re.IGNORECASE)
    if m:
        return _apply_alias(_norm(m.group(1)))

    m = re.match(r"^FUNDS\s+TRANSFER\s+(?:FROM|TO)\s+(.+)$", r, flags=re.IGNORECASE)
    if m:
        return _apply_alias(_norm(re.split(r"[/]", m.group(1))[0]))

    m = re.match(r"^CLG\s+TO\s+(.+)$", r, flags=re.IGNORECASE)
    if m:
        return _apply_alias(_norm(m.group(1)))

    if re.match(r"^(?:NB|IB|N\d+|\d+)\s*:", up) or up.startswith("FROM CASA") or up.startswith("FROM ACCT"):
        if "RD" in up or "CASA" in up or "ACCT" in up or "TRANSFER" in up or "TO KR" in up or "TO RD" in up:
            return "SELF TRANSFER"

    if up.startswith("RD MATURITY") or up.startswith("RD "):
        return "RD (Recurring Deposit)"
    if up.startswith("FD "):
        return "FD (Fixed Deposit)"

    if "ETAX" in up or up.startswith("IB: ETAX"):
        return "ETAX (Income Tax)"

    if up.startswith("CASH WITHDRAWAL") or up.startswith("ATM") or "CASH WDL" in up:
        return "CASH WITHDRAWAL"

    if up.startswith("INT.PD") or "INT.PD:" in up:
        return "SB INTEREST"

    return _apply_alias(extract_generic(r)) if extract_generic(r) else None


class KotakProfile(Profile):
    name = "kotak"

    @classmethod
    def detect(cls, path, raw_text, raw_df):
        text_up = (raw_text or "").upper()
        if "KOTAK" in text_up or "KKBK" in text_up:
            return True
        if raw_df is None:
            return False
        for i in range(min(30, len(raw_df))):
            row = [str(v).strip().lower() for v in raw_df.iloc[i].tolist()]
            if ("particulars" in row and "balance" in row
                    and ("dr" in row and "cr" in row) and "id" in row):
                return True
        return False

    @classmethod
    def parse(cls, path, header_row=None):
        _, raw = _load_raw(path)
        if raw is None:
            raise RuntimeError("Kotak: couldn't load file as a table")
        return _parse_kotak(raw, cls.extract, header_row=header_row)

    @classmethod
    def extract(cls, remarks):
        return _kotak_extract(remarks)


def _find_kotak_header(raw: pd.DataFrame) -> Optional[int]:
    for i in range(min(40, len(raw))):
        row = [str(v).strip().lower() for v in raw.iloc[i].tolist()]
        if "particulars" in row and "balance" in row and "dr" in row and "cr" in row:
            return i
    return None


def _parse_kotak(raw: pd.DataFrame, extract_fn, header_row: Optional[int] = None) -> pd.DataFrame:
    """Kotak statements wrap one logical transaction across TWO physical rows:
       row 1: date | particulars(part 1) | ID | DR | CR | Balance
       row 2: time | particulars(part 2) |    |    |    |
    We locate columns from the header, then merge each start row with the
    continuation row(s) until the next row that carries a DR/CR amount."""
    hdr = header_row if header_row is not None else _find_kotak_header(raw)
    if hdr is None:
        raise RuntimeError("Kotak: couldn't locate header row (Date/Particulars/DR/CR/Balance)")

    header = [str(v).strip().lower() for v in raw.iloc[hdr].tolist()]

    def col_of(*names):
        for idx, h in enumerate(header):
            if h in names:
                return idx
        return None

    c_date = col_of("date")
    c_part = col_of("particulars", "narration", "description")
    c_dr = col_of("dr", "debit", "withdrawal")
    c_cr = col_of("cr", "credit", "deposit")
    c_bal = col_of("balance")

    if c_part is None or c_dr is None or c_cr is None:
        raise RuntimeError("Kotak: required columns (Particulars/DR/CR) not found in header")

    body = raw.iloc[hdr + 1:].reset_index(drop=True)

    def has_amt(i):
        for c in (c_dr, c_cr):
            v = body.iloc[i, c]
            if pd.notna(v) and str(v).strip() not in ("", "nan"):
                return True
        return False

    def cell(i, c):
        if c is None:
            return ""
        v = body.iloc[i, c]
        return "" if pd.isna(v) else str(v).strip()

    records = []
    n = len(body)
    i = 0
    while i < n and not has_amt(i):
        i += 1
    while i < n:
        start = i
        j = i + 1
        while j < n and not has_amt(j):
            j += 1
        part = "".join(cell(k, c_part) for k in range(start, j))
        date = cell(start, c_date)
        dr_raw = cell(start, c_dr)
        cr_raw = cell(start, c_cr)
        bal_raw = cell(start, c_bal)
        records.append({
            "txn_date": date,
            "remarks": part,
            "_dr": dr_raw,
            "_cr": cr_raw,
            "_bal": bal_raw,
        })
        i = j

    df = pd.DataFrame.from_records(records)
    if df.empty:
        return pd.DataFrame(columns=["txn_date", "remarks", "withdrawal",
                                     "deposit", "balance", "counterparty"])

    df["withdrawal"] = _numify(df["_dr"]).abs()
    df["deposit"] = _numify(df["_cr"]).abs()
    df["balance"] = pd.to_numeric(
        df["_bal"].astype(str).str.replace(",", "", regex=False).str.strip(),
        errors="coerce",
    )
    df = df.drop(columns=["_dr", "_cr", "_bal"])
    df = df[(df["withdrawal"] > 0) | (df["deposit"] > 0)].reset_index(drop=True)
    df["counterparty"] = df["remarks"].map(extract_fn)
    return df


# ---------- Generic fallback ----------

class GenericProfile(Profile):
    name = "generic"

    @classmethod
    def detect(cls, path, raw_text, raw_df):
        return True  # always last in registry

    @classmethod
    def parse(cls, path, header_row=None):
        _, raw = _load_raw(path)
        if raw is not None:
            return _parse_from_df(raw, cls.extract, header_row=header_row)
        # File didn't load as Excel/HTML — try tab/comma-separated text.
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            text = f.read()
        sep = "\t" if text.count("\t") > text.count(",") else ","
        rows = list(csv.reader(io.StringIO(text), delimiter=sep))
        df = pd.DataFrame(rows)
        return _parse_from_df(df, cls.extract, header_row=header_row)

    @classmethod
    def extract(cls, remarks):
        n = extract_generic(remarks)
        return _apply_alias(n) if n else None


# Order matters: more-specific profiles first; Generic always last.
PROFILES = [HDFCProfile, AxisProfile, SBIProfile, ICICIProfile, KotakProfile, GenericProfile]


def detect_profile(path: Path) -> type[Profile]:
    raw_text, raw_df = _load_raw(path)
    for prof in PROFILES:
        try:
            if prof.detect(path, raw_text, raw_df):
                return prof
        except Exception:
            continue
    return GenericProfile


# ============================================================
# Writer (unchanged from v1)
# ============================================================

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", start_color="305496")
GROUP_FILL = PatternFill("solid", start_color="D9E1F2")
SUBTOTAL_FILL = PatternFill("solid", start_color="FCE4D6")
GRAND_FILL = PatternFill("solid", start_color="FFE699")

HDR_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
GROUP_FONT = Font(name="Calibri", size=11, bold=True, color="1F3864")
BOLD = Font(name="Calibri", size=11, bold=True)
REG = Font(name="Calibri", size=11)


def _write_sheet(ws, title, df, amount_col, threshold=2):
    raw_keys = [k for k in df["counterparty"] if k]
    canon_map = consolidate(raw_keys)
    df = df.copy()
    df["bucket"] = df["counterparty"].map(lambda k: canon_map.get(k) if k else None)
    df["bucket"] = df["bucket"].fillna("UNSORTED")

    counts = df["bucket"].value_counts()
    named = [b for b in counts.index if b != "UNSORTED" and counts[b] >= threshold]
    totals = df.groupby("bucket")[amount_col].sum()
    named.sort(key=lambda b: -totals.get(b, 0))

    other_df = df[~df["bucket"].isin(named)].copy()
    other_df["bucket"] = "OTHER"

    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F3864")
    ws.merge_cells("A1:E1")

    for i, h in enumerate(["Date", "Counterparty (raw)", "Remarks", "Amount", "Balance"], start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    row = 4

    def write_group(label, sub):
        nonlocal row
        gc = ws.cell(row=row, column=1, value=label)
        gc.font = GROUP_FONT
        gc.fill = GROUP_FILL
        gc.border = BORDER
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = GROUP_FILL
            ws.cell(row=row, column=col).border = BORDER
        row += 1
        first = row
        for _, r in sub.iterrows():
            date_val = r["txn_date"]
            date_cell = ws.cell(row=row, column=1)
            parsed_date = pd.to_datetime(date_val, errors="coerce", dayfirst=True)
            if pd.notna(parsed_date):
                date_cell.value = parsed_date.to_pydatetime()
                date_cell.number_format = "DD-MM-YYYY"
            else:
                date_cell.value = str(date_val) if pd.notna(date_val) else ""
            date_cell.font = REG
            ws.cell(row=row, column=2, value=r["counterparty"] or "").font = REG
            ws.cell(row=row, column=3, value=str(r["remarks"])).font = REG
            ws.cell(row=row, column=4, value=float(r[amount_col])).font = REG
            ws.cell(row=row, column=4).number_format = "#,##0.00"
            bal = r["balance"]
            if pd.notna(bal):
                ws.cell(row=row, column=5, value=float(bal)).font = REG
                ws.cell(row=row, column=5).number_format = "#,##0.00"
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = BORDER
            row += 1
        last = row - 1
        st = ws.cell(row=row, column=3, value="Subtotal")
        st.font = BOLD
        st.alignment = Alignment(horizontal="right")
        st.fill = SUBTOTAL_FILL
        amt = ws.cell(row=row, column=4, value=f"=SUM(D{first}:D{last})")
        amt.font = BOLD
        amt.number_format = "#,##0.00"
        amt.fill = SUBTOTAL_FILL
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = SUBTOTAL_FILL
            ws.cell(row=row, column=col).border = BORDER
        row += 1
        row += 1  # spacer

    def _date_sort(sub_df):
        return sub_df.sort_values(
            "txn_date",
            key=lambda s: pd.to_datetime(s, errors="coerce", dayfirst=True),
            kind="stable",
            na_position="last",
        )

    for b in named:
        write_group(b, _date_sort(df[df["bucket"] == b]))
    if not other_df.empty:
        write_group("OTHER (one-offs)", _date_sort(other_df))

    gt_label = ws.cell(row=row, column=3, value="GRAND TOTAL")
    gt_label.font = BOLD
    gt_label.alignment = Alignment(horizontal="right")
    gt_label.fill = GRAND_FILL
    gt_amt = ws.cell(row=row, column=4, value=f"=SUM(D4:D{row-1})/2")
    gt_amt.font = BOLD
    gt_amt.number_format = "#,##0.00"
    gt_amt.fill = GRAND_FILL
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = GRAND_FILL
        ws.cell(row=row, column=col).border = BORDER

    for i, w in enumerate([12, 38, 60, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def write_workbook(df, out_path, bank_name=""):
    wb = Workbook()
    wb.remove(wb.active)
    title_suffix = f" — {bank_name.upper()}" if bank_name else ""

    dep_df = df[df["deposit"] > 0].copy()
    wd_df = df[df["withdrawal"] > 0].copy()

    ws_dep = wb.create_sheet("Deposits")
    _write_sheet(ws_dep, f"Deposits — grouped by counterparty{title_suffix}", dep_df, "deposit")

    ws_wd = wb.create_sheet("Withdrawals")
    _write_sheet(ws_wd, f"Withdrawals — grouped by counterparty{title_suffix}", wd_df, "withdrawal")

    wb.save(out_path)


# ============================================================
# CLI
# ============================================================

def main():
    if len(sys.argv) < 2:
        print("Usage: python sort_statement.py <statement_file> [output.xlsx]")
        sys.exit(1)
    inp = Path(sys.argv[1]).expanduser()
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else inp.with_name(inp.stem + "_sorted.xlsx")

    prof = detect_profile(inp)
    print(f"Detected bank profile: {prof.name}")

    df = prof.parse(inp)
    print(f"Parsed {len(df)} rows.")
    extracted = df["counterparty"].notna().sum()
    print(f"Counterparty extracted: {extracted}/{len(df)} ({100*extracted/max(len(df),1):.0f}%)")

    write_workbook(df, out, bank_name=prof.name)
    print(f"Wrote: {out}")


if __name__ == "__main__":
    main()
