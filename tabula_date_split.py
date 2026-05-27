"""
Repair Tabula-extracted bank-statement spreadsheets where the Date column has
collapsed into the Narration/column-A cell.

Two failure modes are handled:

  TYPE-2  "date merged only":
      Col A holds  "04/04/25 UPI-ZERODHA BROKING"   (date + first narration line)
      Cols C..G (Chq/Ref, Value Dt, Withdrawal, Deposit, Closing) are still correct.
      Fix: split A -> (Date, Narration). Leave C..G alone.

  TYPE-1  "date merged + whole row shifted left by one":
      Col A holds  "06/04/25 UPI-SWIGGY-..."        (date + narration)
      Col B holds the Chq/Ref number, C holds the Value Dt, D the Withdrawal, etc.
      Fix: split A -> (Date, Narration), then shift B->C, C->D, D->E, E->F, F->G.

Continuation lines (wrapped narration with no leading date) are moved A -> B.
Clean rows (col A already a real date) and blank rows are passed through untouched.

A running-balance reconciliation flags rows where the closing balance does not
follow from the previous closing balance -- these are almost always transactions
that Tabula dropped entirely during PDF extraction, NOT split errors, and must be
recovered manually against the source PDF.

Usage (CLI):
    python tabula_date_split.py input.xlsx output.xlsx

Usage (pipeline):
    from tabula_date_split import clean_statement
    rows, flags = clean_statement("input.xlsx")          # -> list[list], dict
    clean_statement("input.xlsx", out_path="output.xlsx")  # also writes xlsx
"""

from __future__ import annotations

import re
import sys
import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- configuration -----------------------------------------------------------

DATE_FORMATS = (
    "%d/%m/%y", "%d/%m/%Y",
    "%d-%m-%y", "%d-%m-%Y",
    "%d.%m.%y", "%d.%m.%Y",
)

LEAD_DATE_RE = re.compile(
    r"^\s*(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4})\s+(.*)$", re.S
)

OUTPUT_HEADERS = [
    "Date", "Narration", "Chq./Ref.No.", "Value Dt",
    "Withdrawal Amt.", "Deposit Amt.", "Closing Balance", "Check",
]


# --- core helpers ------------------------------------------------------------

def parse_date(token: str) -> Optional[datetime.datetime]:
    """Return a datetime for a dd/mm/yy-ish token, or None if it doesn't parse."""
    token = token.strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.datetime.strptime(token, fmt)
        except ValueError:
            continue
    return None


def _is_shifted_row(b, c) -> bool:
    """
    Decide whether a date-merged row is also shifted left by one.

    In a shift-by-one the whole row moves right by one column, so the Value-Dt
    date lands in C. The presence of a real date in C is the reliable signal:
    on an unshifted (type-2) row C holds the Chq/Ref, which is never a date.
    We deliberately do NOT test the shape of B -- references can be alphanumeric
    (e.g. IMPS "MB08095458841T46") as well as pure digits, so a B-shape test
    would mis-classify those rows. B simply has to be occupied.
    """
    if b in (None, ""):
        return False
    return isinstance(c, datetime.datetime)


# --- main transform ----------------------------------------------------------

def split_rows(ws) -> tuple[list[list], dict[str, int]]:
    """
    Walk the worksheet and return (rows_out, stats).
    rows_out is a list of 7-element lists: Date, Narration, Chq/Ref, Value Dt,
    Withdrawal, Deposit, Closing Balance.
    """
    rows_out: list[list] = []
    stats = {"type1": 0, "type2": 0, "continuation": 0, "clean": 0, "blank": 0}

    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 8)]
        a, b, c, d, e, f, g = vals

        if isinstance(a, datetime.datetime):
            rows_out.append(vals)
            stats["clean"] += 1
            continue

        if a is None:
            if all(v is None for v in vals):
                stats["blank"] += 1
            else:
                stats["continuation"] += 1
            rows_out.append(vals)
            continue

        if isinstance(a, str):
            m = LEAD_DATE_RE.match(a)
            dt = parse_date(m.group(1)) if m else None
            if dt:
                narration = m.group(2).strip()
                if _is_shifted_row(b, c):
                    rows_out.append([dt, narration, b, c, d, e, f])
                    stats["type1"] += 1
                else:
                    rows_out.append([dt, narration, c, d, e, f, g])
                    stats["type2"] += 1
            else:
                rows_out.append([None, a, None, None, None, None, None])
                stats["continuation"] += 1
            continue

        if isinstance(a, int):
            rows_out.append([None, str(a), None, None, None, None, None])
            stats["continuation"] += 1
            continue

        rows_out.append(vals)
        stats["clean"] += 1

    return rows_out, stats


def reconcile_balances(rows_out: list[list], tol: float = 0.01) -> dict[int, str]:
    """
    Check that each transaction's closing balance follows from the previous one:
        prev_closing - withdrawal + deposit == closing
    Returns {row_index_in_rows_out: message} for rows that break the chain.
    """
    flags: dict[int, str] = {}
    prev = None
    for i, row in enumerate(rows_out):
        dt, narr, cref, vdt, wd, dep, cb = row
        if isinstance(dt, datetime.datetime) and isinstance(cb, (int, float)):
            w = wd if isinstance(wd, (int, float)) else 0
            dp = dep if isinstance(dep, (int, float)) else 0
            if prev is not None:
                expected = round(prev - w + dp, 2)
                if abs(expected - round(cb, 2)) > tol:
                    flags[i] = (
                        f"Balance break: {prev} - {w} + {dp} = {expected} "
                        f"!= {cb}. Likely a transaction dropped by Tabula "
                        f"above this row; check the source PDF."
                    )
            prev = cb
    return flags


# --- output ------------------------------------------------------------------

def write_xlsx(rows_out: list[list], flags: dict[int, str], out_path: str,
               font_name: str = "Arial", font_size: int = 11) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cleaned"
    ws.append(OUTPUT_HEADERS)

    base_font = Font(name=font_name, size=font_size)
    head_font = Font(name=font_name, size=font_size, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="305496")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, 9):
        cell = ws.cell(1, col)
        cell.font = head_font
        cell.fill = head_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, row in enumerate(rows_out):
        ws.append(list(row) + [flags.get(i, "")])
        rr = ws.max_row
        warn = i in flags
        for col in range(1, 9):
            cell = ws.cell(rr, col)
            cell.font = base_font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (2, 8)))
            if col in (1, 4) and isinstance(cell.value, datetime.datetime):
                cell.number_format = "dd/mm/yyyy"
            if col in (5, 6, 7) and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
            if warn:
                cell.fill = warn_fill

    widths = {"A": 11, "B": 46, "C": 16, "D": 11, "E": 15, "F": 14, "G": 16, "H": 52}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    wb.save(out_path)


# --- top-level entry point ---------------------------------------------------

def clean_statement(in_path: str, out_path: Optional[str] = None,
                    sheet: Optional[str] = None):
    """
    Clean a Tabula-extracted statement.

    Returns (rows_out, flags, stats). If out_path is given, also writes a formatted xlsx.
    """
    wb = openpyxl.load_workbook(in_path, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    rows_out, stats = split_rows(ws)
    flags = reconcile_balances(rows_out)

    if out_path:
        write_xlsx(rows_out, flags, out_path)

    return rows_out, flags, stats


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python tabula_date_split.py <input.xlsx> [output.xlsx]")
        sys.exit(1)
    src = sys.argv[1]
    dst = sys.argv[2] if len(sys.argv) > 2 else None
    rows_out, flags, stats = clean_statement(src, dst)
    print("Reclassification:", stats)
    print(f"Rows out: {len(rows_out)} | Balance breaks flagged: {len(flags)}")
    for i in sorted(flags):
        narr = rows_out[i][1] or ""
        print(f"  Excel row {i + 2}: {str(narr)[:45]}")
    if dst:
        print(f"Wrote: {dst}")
